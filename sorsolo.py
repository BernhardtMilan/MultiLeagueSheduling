import os
import pandas as pd
import numpy as np
import random
import time
import pickle
from itertools import combinations
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from schedule_validator import ScheduleValidator, parse_occupied_rule #need to fix circle import add refactor somewhere else
from init import *
from league_devider import devide_leagues

# Define the directory where the Excel files are located
#directory = './tablak/'

DAY_INDEX = {day: i for i, day in enumerate(days_of_week)}
SLOT_INDEX = {slot: i for i, slot in enumerate(time_slots)}
WEEK_INDEX = {w: i for i, w in enumerate(weeks)}

def initialize_empty_draw_structure():
    """
    Initializes the data structure for the final draw,
    4 pitches per time slot, all time slots and pitches are initialized as empty.
    """

    schedule = {}

    for week_key in weeks:
        schedule[week_key] = {}

        for day in days_of_week:
            schedule[week_key][day] = {}

            for time_slot in time_slots:
                # Initialize 4 pitches per time slot, each pitch starts empty
                schedule[week_key][day][time_slot] = {
                    1: "", 2: "", 3: "", 4: ""
                }

    return schedule

def add_occupied_times(draw_structure):
    pitches = [1, 2, 3, 4]

    for rule in occupied_rules:
        slots_to_occupy = parse_occupied_rule(rule, days_of_week, time_slots, pitches)
        for (week, day, time, pitch) in slots_to_occupy:
            if week in draw_structure and day in draw_structure[week] and time in draw_structure[week][day]:
                draw_structure[week][day][time][pitch] = "OCCUPIED TIMESLOT"
            else:
                print(f"Invalid slot location in rule: {rule}")

    return draw_structure

# Function to process a single Excel file
def process_excel_file(excel_file):

    allowed_values = {-2, 0, 1, 2}

    try:
        df = pd.read_excel(excel_file, header=None)

        # Check if the league is correctly formatted and is an integer between 1 and 5
        try:
            league = int(df.iloc[0, 1])
            if league not in range(1, 6):
                print(f"Error: Invalid league {league} in file {excel_file}. Must be between 1 and 5.")
                return None, None, None
        except (ValueError, IndexError):
            print(f"Error: Invalid or missing league in file {excel_file}")
            return None, None, None

        # Check if the team name exists and is a non-empty string
        try:
            team_name = df.iloc[0, 3]
            if not isinstance(team_name, str) or not team_name.strip():
                raise ValueError
        except (ValueError, IndexError):
            print(f"Error: Invalid or missing team name in file {excel_file}")
            return None, None, None

        # Extract the table (from row 5 to row 10 and from columns 2 to 7, which are the colored cells)
        try:
            schedule_df = df.iloc[4:10, 1:7].copy()

            # Manually iterate over the DataFrame to validate the values
            for row_idx in range(len(schedule_df)):
                for col_idx in range(1, len(schedule_df.columns)):  # Start at 1 to skip the 'Time' column
                    cell_value = schedule_df.iat[row_idx, col_idx]
                    # Validate that the cell is numeric and within allowed values, or allow NaN
                    if pd.notna(cell_value) and not (isinstance(cell_value, (int, float)) and cell_value in allowed_values):
                        print(f"Error: Invalid value {cell_value} in file {excel_file} at row {row_idx+5}, column {col_idx+1}")
                        return None, None, None
        except Exception as e:
            print(f"Error: Issue reading or validating the schedule table in file {excel_file}")
            return None, None, None

        # Rename the columns based on the days of the week
        schedule_df.columns = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

        # Initialize a dictionary to store the schedule for this team
        schedule = {}

        # Populate the schedule dictionary for this team (skip the "Time" column)
        for i, day in enumerate(days_of_week, start=1):
            schedule[day] = {}
            for j in range(len(schedule_df)):
                time_slot = schedule_df.iloc[j, 0]  # The time (first column)
                availability = schedule_df.iloc[j, i]  # The availability for that day and time
                schedule[day][time_slot] = availability

        return league, team_name, schedule

    except Exception as e:
        print(f"Error: Failed to process file {excel_file}: {e}")
        return None, None, None

def get_input_data_and_sort_to_leagues(directory, plot):
    leagues = [{}, {}, {}, {}, {}]
    team_schedules = {}

    # Get the full paths of all Excel files in the directory
    excel_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.xlsx')]
    # Iterate through each Excel file, process it, and assign the team to the appropriate league
    for excel_file in excel_files:
        league, team_name, schedule = process_excel_file(excel_file)

        # Only assign the team if no errors occurred and the schedule was properly processed
        if league and team_name and schedule:
            if 1 <= league <= 5:
                leagues[league-1][team_name] = schedule

                # Convert to NumPy array (5 days × 6 slots)
                team_array = np.empty((5, 6), dtype=np.int64)

                for day, times in schedule.items():
                    day_idx = DAY_INDEX[day]
                    for slot, value in times.items():
                        slot_idx = SLOT_INDEX[slot]
                        team_array[day_idx, slot_idx] = value

                team_schedules[team_name] = team_array
            else:
                print(f"Error: Unknown league {league} for team {team_name}")

    for i in range(5):
        print(f"Teams in League {i+1}: {len(leagues[i])}")
        print(leagues[i].keys())
        print("")

    # Save variables to file, so we dont have to import every time
    with open(os.path.join(directory, "processed_data.pkl"), "wb") as f:
        pickle.dump((leagues, team_schedules), f)

    return leagues, team_schedules

def get_input_data_from_saves(directory, plot=False):
    import pickle

    filepath = os.path.join(directory, "processed_data.pkl")
    try:
        with open(filepath, "rb") as f:
            leagues, team_schedules = pickle.load(f)

        if plot:
            for i in range(5):
                print(f"Teams in League {i+1}: {len(leagues[i])}")
                print(leagues[i].keys())
                print("")

        return leagues, team_schedules
    except FileNotFoundError:
        print(f"No saved data found in {filepath}. Run get_input_data_and_sort_to_leagues first.")
        return None, None

def visualize_devided_leagues(devided_leagues, division_counts):
    print("--------------------------DIVIDED LEAGUES-----------------------------------")
    # Print teams in each divided league with the original league and subdivision number
    original_league_number = 1  # Tracks the original league number
    subdivision_index = 1       # Tracks the subdivision within each league

    for i in range(len(devided_leagues)):
        print(f"Teams in League {original_league_number}/{subdivision_index}: {len(devided_leagues[i].keys())}")
        print(devided_leagues[i].keys())
        print("")

        # Check if we need to move to the next original league or just increment the subdivision
        if subdivision_index < division_counts[original_league_number - 1]:
            subdivision_index += 1  # Move to the next subdivision within the current league
        else:
            original_league_number += 1  # Move to the next original league
            subdivision_index = 1        # Reset subdivision index for the new league

def create_all_pairs(leagues):
    leagues_all_games = [[] for i in range(len(leagues))]
    league_num = 0
    for league in leagues:
        team_names = list(league.keys())
    
        # Generate all unique combinations (pairings) of teams
        leagues_all_games[league_num] = list(combinations(team_names, 2))
        league_num += 1
    return leagues_all_games

def visualize_league_games(leagues_all_games, division_counts):
    # Reset original league and subdivision indices for printing the games
    original_league_number = 1
    subdivision_index = 1

    for i in range(len(leagues_all_games)):
        league_label = f"{original_league_number}/{subdivision_index}"
        if i == 0:
            print(f"{league_label}. league games:")
            print(leagues_all_games[i])
        else:
            print(f"Sample from league {league_label}:")
            print(leagues_all_games[i][2])  # Print a sample of games

        # Update league and subdivision indices similarly as above
        if subdivision_index < division_counts[original_league_number - 1]:
            subdivision_index += 1  # Move to the next subdivision within the current league
        else:
            original_league_number += 1  # Move to the next original league
            subdivision_index = 1        # Reset subdivision index for the new league

def sort_random_matches(draw_structure, leagues_all_games):
    for league in leagues_all_games:
        for game in league:
            placed = False
            
            while not placed:
                # Randomly select a week, day, and time slot
                week_key = random.choice(list(draw_structure.keys()))
                day_key = random.choice(list(draw_structure[week_key].keys()))
                time_slot_key = random.choice(list(draw_structure[week_key][day_key].keys()))
                
                # Check each pitch in the selected time slot for availability
                for pitch, match in draw_structure[week_key][day_key][time_slot_key].items():
                    if match == '':
                        # Place the game and mark it as placed
                        draw_structure[week_key][day_key][time_slot_key][pitch] = game
                        placed = True
                        break  # Stop checking pitches once we place the game
                        
    return draw_structure

def calulate_team_metric(draw_structure, team_schedules, league_teams):

    data = defaultdict(lambda: {
        "availability": 0,
        "match_bunching_penalty": 0,
        "idle_gap_penalty": 0,
        "spread_reward": 0,
        "number_of_matches": 0
    })

    matches = flatten_matches(draw_structure)
    team_week_counts = defaultdict(lambda: [0] * number_of_weeks)

    for week_idx, day_key, timeslot_key, pitch, (team1, team2) in matches:
        day_idx  = DAY_INDEX[day_key]
        slot_idx = SLOT_INDEX[timeslot_key]

        for team in (team1, team2):
            data[team]["availability"] += int(team_schedules[team][day_idx, slot_idx])
            data[team]["number_of_matches"] += 1
            team_week_counts[team][week_idx - 1] += 1
        
    for team, counts in team_week_counts.items():
        # bunching: playing more than once in a week
        data[team]["match_bunching_penalty"] = sum((c - 1) for c in counts if c > 1)

        # spread + idle gaps
        weeks_played = [i for i, c in enumerate(counts) if c > 0]
        data[team]["spread_reward"] = len(weeks_played)

        if len(weeks_played) > 1:
            gaps = [weeks_played[i+1] - weeks_played[i] - 1 for i in range(len(weeks_played)-1)]
            # penalize gaps of 2+ idle weeks (same rule as your global metric)
            data[team]["idle_gap_penalty"] = sum(1 for g in gaps if g >= 2)

        data[team]["availability"] *= weights["availability"]
        data[team]["match_bunching_penalty"] *= weights["match_bunching_penalty"]
        data[team]["idle_gap_penalty"] *= weights["idle_gap_penalty"]
        data[team]["spread_reward"] *= weights["spread_reward"]

    return dict(data)

def flatten_matches(draw_structure):
    matches = []
    for week_idx, week in enumerate(draw_structure.values(), start=1):
        for day_key, day in week.items():
            for timeslot_key, timeslot in day.items():
                for pitch, match in timeslot.items():
                    if match and match != "OCCUPIED TIMESLOT":
                        matches.append((week_idx, day_key, timeslot_key, pitch, match))
    return matches

def calculate_metric(draw_structure, team_schedules, league_teams):

    #start_time = time.time()

    total_availability_score = 0

    team_week_counts = defaultdict(lambda: [0] * number_of_weeks)

    matches = flatten_matches(draw_structure)

    L1_teams = set(league_teams["L1"])
    L1_pitch_penalty = 0

    for week_idx, day_key, timeslot_key, pitch, (team1, team2) in matches:
        team1_schedule = team_schedules[team1]
        team2_schedule = team_schedules[team2]

        day_idx = DAY_INDEX[day_key]
        slot_idx = SLOT_INDEX[timeslot_key]

        team1_availability = team1_schedule[day_idx, slot_idx]
        team2_availability = team2_schedule[day_idx, slot_idx]

        total_availability_score += team1_availability + team2_availability

        team_week_counts[team1][week_idx - 1] += 1
        team_week_counts[team2][week_idx - 1] += 1

        # L1 pitch penalty
        if team1 in L1_teams and pitch != 1:
            L1_pitch_penalty += 1

    bunching_penalty = 0
    idle_gap_penalty = 0
    spread_reward = 0

    for counts in team_week_counts.values():
        weeks_played = [i for i, count in enumerate(counts) if count > 0]

        bunching_penalty += sum(count - 1 for count in counts if count > 1)

        if len(weeks_played) > 1:
            gaps = [weeks_played[i+1] - weeks_played[i] - 1 for i in range(len(weeks_played)-1)]
            idle_gap_penalty += sum(1 for gap in gaps if gap >= 2)

        spread_reward += len(weeks_played)

    total_metric = (
        weights["availability"] * total_availability_score
        + weights["match_bunching_penalty"] * bunching_penalty
        + weights["idle_gap_penalty"] * idle_gap_penalty
        + weights["spread_reward"] * spread_reward
        + weights["L1_pitch_penalty"] * L1_pitch_penalty
    )

    #elapsed = time.time() - start_time
    #print(f"Time taken for calculating metric: {elapsed:.5f} seconds")

    return total_metric, team_week_counts

def calculate_final_metric(draw_structure, team_schedules, league_teams):

    total_availability_score = 0
    number_of_matches = 0
    value_counts = [0, 0, 0, 0]
    value_map = {-2: 0, 0: 1, 1: 2, 2: 3}

    team_weeks = defaultdict(list)

    L1_teams = set(league_teams["L1"])
    L1_pitch_penalty = 0

    for week_idx, (week_name, week) in enumerate(draw_structure.items(), start=1):
        for day_key, day in week.items():
            for timeslot_key, timeslot in day.items():
                for pitch, match in timeslot.items():
                    if match and match != "OCCUPIED TIMESLOT":
                        team1, team2 = match
                        team1_schedule = team_schedules[team1]
                        team2_schedule = team_schedules[team2]

                        # If a team schedule is not found, skip this match
                        if team1_schedule is None or team2_schedule is None:
                            print(f"Warning: Schedule not found for {team1} or {team2}")
                            continue

                        day_idx = DAY_INDEX[day_key]
                        slot_idx = SLOT_INDEX[timeslot_key]

                        try:
                            team1_availability = int(team_schedules[team1][day_idx, slot_idx])
                            team2_availability = int(team_schedules[team2][day_idx, slot_idx])
                        except KeyError:
                            # If availability data is missing for a team or timeslot, skip this match
                            print(f"Warning: Missing data for {team1} or {team2} at {day_key} {timeslot_key}")
                            continue

                        match_metric = team1_availability + team2_availability
                        total_availability_score += match_metric
                        number_of_matches += 1

                        for value in (team1_availability, team2_availability):
                            if value in value_map:
                                value_counts[value_map[value]] += 1

                        team_weeks[team1].append(week_idx)
                        team_weeks[team2].append(week_idx)

                        if team1 in L1_teams and pitch != 1:
                            L1_pitch_penalty += 1

    bunching_penalty = 0
    idle_gap_penalty = 0
    spread_reward = 0

    for team, weeks_played in team_weeks.items():
        weeks_list = sorted(weeks_played)
        week_counts = defaultdict(int)

        # Count how many matches per week
        # Penalty: too many matches in the same week
        for w in weeks_played:
            week_counts[w] += 1
        for week in week_counts:
            if week_counts[week] > 1:
                bunching_penalty += (week_counts[week] - 1)

        # Penalty: long gaps, not including before the first and after the last week
        gaps = [
            weeks_list[i+1] - weeks_list[i] - 1
            for i in range(len(weeks_list)-1)
        ]
        idle_gap_penalty += sum(1 for gap in gaps if gap >= 2)

        # Reward: playing in many different weeks
        spread_reward += len(set(weeks_list))

    total_metric = (
        weights["availability"] * total_availability_score
        + weights["match_bunching_penalty"] * bunching_penalty
        + weights["idle_gap_penalty"] * idle_gap_penalty
        + weights["spread_reward"] * spread_reward
        + weights["L1_pitch_penalty"] * L1_pitch_penalty
    )

    scores = [total_availability_score, bunching_penalty, idle_gap_penalty, spread_reward, L1_pitch_penalty]

    return total_metric, scores, number_of_matches, value_counts

def calculate_possible_max_metric(leagues_all_games):

    total_teams = 0
    total_matches = 0
    total_availability_score = 0
    total_spread_reward = 0

    for league_matches in leagues_all_games:
        teams = set()
        for match in league_matches:
            teams.update(match)

        num_teams = len(teams)
        num_matches = len(league_matches)

        total_teams += num_teams
        total_matches += num_matches

        total_availability_score += num_matches * 4

        total_spread_reward += num_teams * (num_teams - 1)

    max_metric = (
        weights["availability"] * total_availability_score +
        weights["spread_reward"] * total_spread_reward
    )

    return max_metric

def build_league_team_map(devided_leagues, league_names):
    league_team_map = {}

    for league_dict, league_name in zip(devided_leagues, league_names):
        teams = list(league_dict.keys())
        league_team_map[league_name] = teams

    return league_team_map

def generate_output(draw_structure, league_teams, filename):
    day_labels = {
        'Monday': 'Hétfő',
        'Tuesday': 'Kedd',
        'Wednesday': 'Szerda',
        'Thursday': 'Csütörtök',
        'Friday': 'Péntek'
    }
    time_labels = {
        '17:00-18:00': '17-18',
        '18:00-19:00': '18-19',
        '19:00-20:00': '19-20',
        '20:00-21:00': '20-21',
        '21:00-22:00': '21-22',
        '22:00-23:00': '22-23'
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    row_offset = 1
    for week_index, (week_key, week_data) in enumerate(draw_structure.items(), start=1):
        # Week label in column A
        ws.cell(row=row_offset, column=1, value=f"{week_index}.hét").font = Font(bold=True)

        # Merge cells and set day headers (across 4 pitches)
        for d_index, day in enumerate(days_of_week):
            start_col = 2 + d_index * 4
            end_col = start_col + 3
            ws.merge_cells(start_row=row_offset, start_column=start_col, end_row=row_offset, end_column=end_col)
            header_cell = ws.cell(row=row_offset, column=start_col)
            header_cell.value = day_labels[day]
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.font = Font(bold=True)

        # Set pitch headers
        for d_index in range(len(days_of_week)):
            for p_index in range(4):
                col = 2 + d_index * 4 + p_index
                ws.cell(row=row_offset + 1, column=col, value=f"{p_index + 1}.pálya")

        # Fill matches
        for t_index, time in enumerate(time_slots):
            row = row_offset + 2 + t_index
            ws.cell(row=row, column=1, value=time_labels[time])  # time label in first column

            for d_index, day in enumerate(days_of_week):
                for p_index in range(4):
                    col = 2 + d_index * 4 + p_index
                    match = week_data.get(day, {}).get(time, {}).get(p_index + 1, "")
                    if isinstance(match, tuple):
                        ws.cell(row=row, column=col, value=f"{match[0]}-{match[1]}")
                    elif match == "OCCUPIED TIMESLOT":
                        ws.cell(row=row, column=col, value="X")

        # Move offset to next week (1 row for day header + 1 for pitch header + 6 for time slots)
        row_offset += 8
    
    ws2 = wb.create_sheet("Leagues")
    ws2['A1'] = "League"
    ws2['B1'] = "Teams"
    ws2['A1'].font = ws2['B1'].font = Font(bold=True)

    row = 2
    for league, teams in league_teams.items():
        ws2.cell(row=row, column=1, value=league)
        ws2.cell(row=row, column=2, value=", ".join(teams))
        row += 1

    wb.save(filename)
    return

def initial_sort(directory, plot=True):

    print(f"importing data from {directory}...")
    
    draw_structure = initialize_empty_draw_structure()

    draw_structure = add_occupied_times(draw_structure)

    #import fresh from excel files
    #leagues, team_schedules = get_input_data_and_sort_to_leagues(directory, plot)
    # pre saved data, for faster initiakuzation
    leagues, team_schedules = get_input_data_from_saves(directory, plot)

    devided_leagues, division_counts = devide_leagues(leagues, devision_strategy, plot)
    if plot: visualize_devided_leagues(devided_leagues, division_counts)

    league_teams = build_league_team_map(devided_leagues, league_names)

    leagues_all_games = create_all_pairs(devided_leagues)
    if plot: visualize_league_games(leagues_all_games, division_counts)

    draw_structure = sort_random_matches(draw_structure, leagues_all_games)

    metric, scores, number_of_matches, value_counts = calculate_final_metric(draw_structure, team_schedules, league_teams)

    if plot:
        print("")
        print("Samples from draw structure:")
        print("Week 2, Tuesday:")
        print(draw_structure["Week 2"]["Tuesday"])
        print("")
        print("Week 4, Friday:")
        print(draw_structure["Week 4"]["Friday"])
        print("")
        print("Week 10, Monday, 18:00-19:00:")
        print(draw_structure["Week 10"]["Monday"]["18:00-19:00"])
        print("")
        try:
            print("Week 14, Monday, 20:00-21:00:")
            print(draw_structure["Week 14"]["Monday"]["20:00-21:00"])
        except KeyError:
            print("Week 14, Monday, 20:00-21:00: Not available")

        print("")
        print("METRIC:")
        print(metric)
        print("")
        print("Scores:")
        print("[total_availability_score, bunching_penalty, idle_gap_penalty, spread_reward, L1_pitch_penalty]")
        print(scores)
        print("Weighted scores:")
        print(f'[{weights["availability"] * scores[0]}, {weights["match_bunching_penalty"] * scores[1]}, {weights["idle_gap_penalty"] * scores[2]}, {weights["spread_reward"] * scores[3]}, {weights["L1_pitch_penalty"] * scores[4]}]')
        print("")
        print("The number of matches:")
        print(number_of_matches)
        print("All possible matches (16weeks, 5days, 6timeslots, 4pithes) - occupied times:")
        print(16*5*6*4 - 35)

        print("")
        print("Match avaliability:")
        print("[bad, no answer, might, good]")
        print(value_counts)

        print(" ")
        print("Validating schedule...")
    ScheduleValidator(draw_structure, devided_leagues, division_counts)

    sould_save_random_sort = False

    if sould_save_random_sort:
        np.save('random_sort.npy', draw_structure)

    possible_max_metric = calculate_possible_max_metric(leagues_all_games)
    
    return (draw_structure, team_schedules, possible_max_metric, league_teams)

if __name__ == "__main__":
    initial_sort(directory, plot=True)